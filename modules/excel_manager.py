"""
modules/excel_manager.py — Contrôle Excel
==========================================
Semaine 10 — Création, lecture et manipulation de fichiers Excel.

Dépendances :
  pip install openpyxl          # Manipulation .xlsx
  pip install xlsxwriter        # Graphiques avancés (optionnel)

Fonctionnalités :
  - Créer un fichier Excel avec données et mise en forme
  - Lire/extraire des données d'un .xlsx existant
  - Générer des graphiques (barres, lignes, camembert)
  - Créer un tableau de rapport structuré
  - Ajouter des formules (SUM, AVERAGE, etc.)
  - Appliquer des styles (couleurs, gras, largeur colonnes)
"""

from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from config.logger import get_logger

logger = get_logger(__name__)

# ── Imports optionnels ────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl non installé. pip install openpyxl")

DEFAULT_OUTPUT_DIR = Path.home() / "Documents" / "Jarvis"

# Couleurs thème
COLOR_HEADER_BG  = "1F497D"  # Bleu foncé
COLOR_HEADER_FG  = "FFFFFF"  # Blanc
COLOR_ALT_ROW    = "DCE6F1"  # Bleu clair
COLOR_TOTAL_BG   = "FFEB9C"  # Jaune


class ExcelManager:
    """
    Gestionnaire Excel complet via openpyxl.
    Toutes les méthodes retournent { success, message, data }.
    """

    def __init__(self, output_dir: str = None):
        self._output_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
        self._output_dir.mkdir(parents=True, exist_ok=True)

    # ══════════════════════════════════════════════════════════════════════════
    #  CRÉATION
    # ══════════════════════════════════════════════════════════════════════════

    def create_spreadsheet(
        self,
        title: str,
        sheets: list[dict],
        filename: str = None,
        output_dir: str = None,
        open_after: bool = True,
    ) -> dict:
        """
        Crée un fichier Excel avec une ou plusieurs feuilles.

        Args:
            title   : Titre du classeur (affiché dans la propriété)
            sheets  : Liste de feuilles, chacune avec :
                      {
                        "name"    : str — Nom de l'onglet
                        "headers" : [str, ...] — Noms de colonnes
                        "rows"    : [[val, val, ...], ...] — Données
                        "totals"  : bool — Ajouter une ligne de totaux
                        "chart"   : dict — Config graphique (optionnel)
                        "widths"  : [int, ...] — Largeurs colonnes (optionnel)
                      }
            filename   : Nom fichier sans extension
            output_dir : Répertoire de sortie
            open_after : Ouvrir après création

        Returns:
            data : path, name, sheets_count
        """
        if not OPENPYXL_AVAILABLE:
            return self._err("openpyxl non installé. pip install openpyxl")

        out_dir = Path(output_dir) if output_dir else self._output_dir
        out_dir.mkdir(parents=True, exist_ok=True)

        if not filename:
            safe = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')
            filename = f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filepath = out_dir / f"{filename}.xlsx"

        try:
            wb = Workbook()
            wb.properties.title = title
            # Supprimer la feuille par défaut
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            for sheet_def in sheets:
                ws = wb.create_sheet(title=sheet_def.get("name", "Feuille"))
                headers = sheet_def.get("headers", [])
                rows    = sheet_def.get("rows", [])
                widths  = sheet_def.get("widths", [])

                # En-têtes
                if headers:
                    self._write_headers(ws, headers, widths)

                # Données
                start_row = 2 if headers else 1
                for i, row in enumerate(rows):
                    ws_row = start_row + i
                    for j, val in enumerate(row):
                        col = j + 1
                        ws.cell(row=ws_row, column=col, value=val)
                    # Alternance de couleur
                    if i % 2 == 1:
                        self._fill_row(ws, ws_row, len(row), COLOR_ALT_ROW)

                # Ligne de totaux
                if sheet_def.get("totals") and rows and headers:
                    total_row = start_row + len(rows)
                    ws.cell(row=total_row, column=1, value="TOTAL")
                    ws.cell(row=total_row, column=1).font = Font(bold=True)
                    for j in range(1, len(headers)):
                        col_letter = get_column_letter(j + 1)
                        formula = f"=SUM({col_letter}{start_row}:{col_letter}{total_row - 1})"
                        ws.cell(row=total_row, column=j + 1, value=formula)
                    self._fill_row(ws, total_row, len(headers), COLOR_TOTAL_BG, bold=True)

                # Figer la première ligne
                if headers:
                    ws.freeze_panes = "A2"

                # Graphique
                if sheet_def.get("chart") and headers and rows:
                    self._add_chart(
                        ws,
                        sheet_def["chart"],
                        data_start_row=start_row,
                        data_end_row=start_row + len(rows) - 1,
                        n_cols=len(headers),
                        anchor_col=len(headers) + 2,
                    )

            wb.save(str(filepath))
            logger.info(f"Excel créé : {filepath.name}")

            if open_after:
                self._open_file(str(filepath))

            return self._ok(
                f"Fichier Excel '{filepath.name}' créé ({len(sheets)} feuille(s)).",
                {
                    "path":   str(filepath),
                    "name":   filepath.name,
                    "sheets": len(sheets),
                    "opened": open_after,
                }
            )
        except Exception as e:
            logger.error(f"Erreur création Excel : {e}")
            return self._err(f"Impossible de créer le fichier Excel : {e}")

    def create_report(
        self,
        title: str,
        data: list[dict],
        filename: str = None,
        output_dir: str = None,
        include_chart: bool = True,
        open_after: bool = True,
    ) -> dict:
        """
        Génère un rapport Excel complet depuis une liste de dicts.

        Args:
            title         : Titre du rapport
            data          : [{"col1": val, "col2": val, ...}, ...]
            include_chart : Ajouter un graphique automatique
            filename, output_dir, open_after : idem create_spreadsheet
        """
        if not data:
            return self._err("Aucune donnée fournie pour le rapport.")

        headers = list(data[0].keys())
        rows    = [[d.get(h, "") for h in headers] for d in data]

        chart_config = None
        if include_chart and len(headers) >= 2:
            # Graphique barres sur les 2 premières colonnes numériques
            num_cols = [
                i for i, h in enumerate(headers)
                if any(isinstance(r[i], (int, float)) for r in rows)
            ]
            if num_cols:
                chart_config = {
                    "type":    "bar",
                    "title":   title,
                    "x_col":   1,
                    "y_col":   num_cols[0] + 1,
                }

        sheets = [{
            "name":    "Rapport",
            "headers": headers,
            "rows":    rows,
            "totals":  True,
            "chart":   chart_config,
        }]
        return self.create_spreadsheet(
            title=title,
            sheets=sheets,
            filename=filename,
            output_dir=output_dir,
            open_after=open_after,
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  LECTURE
    # ══════════════════════════════════════════════════════════════════════════

    def read_spreadsheet(
        self,
        path: str,
        sheet_name: str = None,
        max_rows: int = 100,
    ) -> dict:
        """
        Lit les données d'un fichier Excel existant.

        Args:
            path       : Chemin du fichier .xlsx
            sheet_name : Nom de l'onglet (défaut : premier onglet)
            max_rows   : Nombre max de lignes à retourner

        Returns:
            data : headers, rows, sheet_name, total_rows, display
        """
        if not OPENPYXL_AVAILABLE:
            return self._err("openpyxl non installé.")

        path_obj = Path(path)
        if not path_obj.exists():
            return self._err(f"Fichier introuvable : '{path}'")
        if path_obj.suffix.lower() not in (".xlsx", ".xlsm", ".xltx"):
            return self._err(f"Format non supporté : '{path_obj.suffix}'. Utilise .xlsx")

        try:
            wb = load_workbook(str(path_obj), data_only=True)

            # Sélectionner la feuille
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Lire toutes les données
            all_rows = list(ws.values)
            if not all_rows:
                return self._ok(
                    f"Feuille '{sheet_name}' vide.",
                    {"headers": [], "rows": [], "sheet_name": sheet_name, "total_rows": 0}
                )

            # Première ligne = en-têtes si présente
            headers   = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = []
            for row in all_rows[1:max_rows + 1]:
                data_rows.append([str(v) if v is not None else "" for v in row])

            total_rows = len(all_rows) - 1

            # Affichage tableau texte
            col_widths = [max(len(h), max((len(str(r[i])) for r in data_rows), default=0))
                          for i, h in enumerate(headers)]
            lines = []
            header_line = " | ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers))
            lines.append(header_line)
            lines.append("─" * len(header_line))
            for row in data_rows[:20]:
                lines.append(" | ".join(str(v).ljust(col_widths[i])
                                         for i, v in enumerate(row)))
            if total_rows > 20:
                lines.append(f"  ... et {total_rows - 20} lignes supplémentaires")

            logger.info(f"Excel lu : {path_obj.name} — {total_rows} lignes, {len(headers)} colonnes")
            return self._ok(
                f"Fichier '{path_obj.name}' lu : {total_rows} ligne(s), {len(headers)} colonne(s).",
                {
                    "headers":    headers,
                    "rows":       data_rows,
                    "sheet_name": sheet_name,
                    "sheets":     wb.sheetnames,
                    "total_rows": total_rows,
                    "total_cols": len(headers),
                    "path":       str(path_obj),
                    "display":    "\n".join(lines),
                }
            )
        except Exception as e:
            logger.error(f"Erreur lecture Excel : {e}")
            return self._err(f"Impossible de lire '{path_obj.name}' : {e}")

    def get_cell(self, path: str, cell: str, sheet_name: str = None) -> dict:
        """
        Retourne la valeur d'une cellule spécifique (ex: "A1", "B3").
        """
        if not OPENPYXL_AVAILABLE:
            return self._err("openpyxl non installé.")
        path_obj = Path(path)
        if not path_obj.exists():
            return self._err(f"Fichier introuvable : '{path}'")
        try:
            wb = load_workbook(str(path_obj), data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            value = ws[cell].value
            return self._ok(
                f"Cellule {cell} = {value}",
                {"cell": cell, "value": value, "sheet": ws.title}
            )
        except Exception as e:
            return self._err(f"Erreur lecture cellule : {e}")

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPERS
    # ══════════════════════════════════════════════════════════════════════════

    def _write_headers(self, ws, headers: list, widths: list = None):
        """Écrit et stylise la ligne d'en-têtes."""
        header_font    = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
        header_fill    = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        header_align   = Alignment(horizontal="center", vertical="center")

        for j, header in enumerate(headers):
            col = j + 1
            cell = ws.cell(row=1, column=col, value=str(header))
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

            # Largeur colonne
            width = (widths[j] if widths and j < len(widths)
                     else max(len(str(header)) + 4, 12))
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 20

    def _fill_row(self, ws, row_num: int, n_cols: int, color: str, bold: bool = False):
        """Applique une couleur de fond à toute une ligne."""
        fill = PatternFill("solid", fgColor=color)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            if bold:
                cell.font = Font(bold=True)

    def _add_chart(
        self,
        ws,
        config: dict,
        data_start_row: int,
        data_end_row: int,
        n_cols: int,
        anchor_col: int,
    ):
        """
        Ajoute un graphique à la feuille.

        config keys : type (bar|line|pie), title, x_col (int), y_col (int)
        """
        chart_type = config.get("type", "bar").lower()
        chart_title = config.get("title", "Graphique")
        x_col = config.get("x_col", 1)
        y_col = config.get("y_col", 2)

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title  = chart_title
        chart.style  = 10
        chart.width  = 15
        chart.height = 10

        # Données Y
        data_ref = Reference(
            ws,
            min_col=y_col, max_col=y_col,
            min_row=data_start_row - 1,  # Inclure l'en-tête
            max_row=data_end_row
        )
        chart.add_data(data_ref, titles_from_data=True)

        # Labels X
        cats = Reference(ws, min_col=x_col, min_row=data_start_row, max_row=data_end_row)
        chart.set_categories(cats)

        # Position
        anchor = f"{get_column_letter(anchor_col)}{data_start_row}"
        ws.add_chart(chart, anchor)

    @staticmethod
    def _open_file(path: str):
        try:
            os.startfile(path)
        except Exception:
            import subprocess, sys
            if sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])

    @staticmethod
    def _ok(message: str, data=None) -> dict:
        return {"success": True,  "message": message, "data": data}

    @staticmethod
    def _err(message: str, data=None) -> dict:
        return {"success": False, "message": message, "data": data}